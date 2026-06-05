#!/usr/bin/env python3
"""Generate team level evaluation Excel draft (0.2 step scale + rubric sub-score sheet).

실행할 때마다 `팀역량레벨평가_초안_YYYYMMDD_HHMMSS.xlsx` 형태의 새 파일만 생성합니다.
기존 xlsx는 덮어쓰지 않습니다.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def new_workbook_path() -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path(__file__).resolve().parent / f"팀역량레벨평가_초안_{stamp}.xlsx"


thin = Side(style="thin", color="CCCCCC")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font = Font(bold=True, size=11)
hdr_fill = PatternFill("solid", fgColor="E2EFDA")
# 사용자 직접 입력(또는 목록 선택) 셀 — 수식 열과 구분
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
# 팀장 시트 C~H: 자체평가에서 참조하는 수식(연한 파랑). 팀장만의 값으로 바꾸려면 해당 셀 수식 삭제 후 입력.
PULL_FILL = PatternFill("solid", fgColor="DDEBF7")
wrap = Alignment(wrap_text=True, vertical="top")


def paint_rect_fill(ws, min_row: int, max_row: int, min_col: int, max_col: int, fill: PatternFill) -> None:
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).fill = fill


def build_fraction_accumulation_sheet(
    ws,
    title: str,
    note: str,
    sample_rows: list[tuple],
    *,
    link_self_sheet: str | None = None,
) -> None:
    """sample_rows: (ym, name, int, d,e,f,g,h). link_self_sheet 설정 시 C~H는 자체평가 시트 동일 키에서 참조."""
    ws.merge_cells("A1:M1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = wrap
    ws.merge_cells("A2:M2")
    ws["A2"] = note
    ws["A2"].alignment = wrap
    ws["A2"].font = Font(size=10, color="444444")

    sh = 3
    shdrs = [
        "평가월",
        "이름",
        "정수레벨",
        "자율성",
        "범위·난이도",
        "협업·영향",
        "품질·완결성",
        "전문성·표준화",
        "누적(우선순위)",
        "캡적용",
        "MROUND0.2",
        "제안_종합",
        "키(자동)",
    ]
    for c, h in enumerate(shdrs, 1):
        cell = ws.cell(sh, c, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border_all
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    yn_dv = DataValidation(type="list", formula1='"충족,미충족"', allow_blank=True)
    yn_dv.add("D4:H500")
    ws.add_data_validation(yn_dv)

    int_dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    int_dv.add("C4:C500")
    ws.add_data_validation(int_dv)

    src = f"'{link_self_sheet}'" if link_self_sheet else None

    for ri, row in enumerate(sample_rows, 4):
        ym, name, lv, d, e, f, g, h = row
        ws.cell(ri, 1, ym)
        ws.cell(ri, 2, name)
        if not link_self_sheet:
            ws.cell(ri, 3, lv)
            ws.cell(ri, 4, d)
            ws.cell(ri, 5, e)
            ws.cell(ri, 6, f)
            ws.cell(ri, 7, g)
            ws.cell(ri, 8, h)

    if link_self_sheet:
        cols_letter = ["C", "D", "E", "F", "G", "H"]
        for ri in range(4, 501):
            for col_letter, col_idx in zip(cols_letter, range(3, 9)):
                ws.cell(
                    ri,
                    col_idx,
                    f'=IF(OR($A{ri}="",$B{ri}=""),"",IFERROR(INDEX({src}!${col_letter}$4:${col_letter}$500,'
                    f'MATCH($A{ri}&"|"&$B{ri},{src}!$M$4:$M$500,0)),""))',
                )

    for ri in range(4, 501):
        ws.cell(
            ri,
            9,
            (
                f'=IF(OR($A{ri}="",$B{ri}="",$C{ri}=""),"",IF(COUNTA($D{ri}:$H{ri})<5,"",'
                f'IF($D{ri}<>"충족",0,0.2)'
                f'+IF(AND($D{ri}="충족",$E{ri}="충족"),0.2,0)'
                f'+IF(AND($D{ri}="충족",$E{ri}="충족",$F{ri}="충족"),0.2,0)'
                f'+IF(AND($D{ri}="충족",$E{ri}="충족",$F{ri}="충족",$G{ri}="충족"),0.2,0)'
                f'+IF(AND($D{ri}="충족",$E{ri}="충족",$F{ri}="충족",$G{ri}="충족",$H{ri}="충족"),0.2,0)))'
            ),
        )
        ws.cell(ri, 10, f'=IF($I{ri}="","",MIN($I{ri},IF($G{ri}="충족",10,0.4)))')
        ws.cell(ri, 11, f'=IF($J{ri}="","",MROUND($J{ri},0.2))')
        ws.cell(ri, 12, f'=IF($K{ri}="","",$C{ri}+$K{ri})')
        ws.cell(ri, 13, f'=IF($A{ri}="","",$A{ri}&"|"&$B{ri})')

    widths = [11, 10, 10, 10, 12, 10, 12, 14, 14, 9, 12, 11, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in range(4, 501):
        for col in (9, 10, 11, 12):
            ws.cell(row, col).number_format = "0.0"

    paint_rect_fill(ws, 4, 500, 1, 2, INPUT_FILL)
    if link_self_sheet:
        paint_rect_fill(ws, 4, 500, 3, 8, PULL_FILL)
    else:
        paint_rect_fill(ws, 4, 500, 3, 8, INPUT_FILL)

    ws.freeze_panes = "A4"


def main() -> None:
    out = new_workbook_path()
    wb = Workbook()

    # --- 코드북 ---
    ws0 = wb.active
    ws0.title = "코드북"

    ws0["A1"] = "팀 역량 레벨 평가 — 코드북 · 집계 규칙"
    ws0["A1"].font = Font(bold=True, size=14)

    ws0["A3"] = "【레벨 정의 (본부)】"
    ws0["A3"].font = hdr_font

    levels = [
        (
            "Level 1 – 기초",
            "신규 입사자 또는 신입 만 1년 미만 수준. 매뉴얼·가이드 없이는 수행 어렵고, 항상 밀착된 지시·검토가 필요.",
        ),
        (
            "Level 2 – 기본",
            "정해진 절차나 템플릿을 따라 기본 업무 수행 가능. 단순·반복 업무는 할 수 있으나, 예외 상황 도움 필요.",
        ),
        (
            "Level 3 – 독립 수행",
            "일반적인 상황에서는 스스로 계획·수행·산출까지 가능. 복잡하거나 전략적인 이슈는 상위자 코칭·리뷰 필요.",
        ),
        (
            "Level 4 – 숙련",
            "대부분의 상황에서 높은 품질로 독립 수행, 복잡한 케이스도 안정적으로 처리. 문제를 미리 예측하고, 업무 방식을 개선하거나 남을 코칭하기 시작.",
        ),
        (
            "Level 5 – 전문가",
            "영역 내에서 레퍼런스 역할, 기준·프레임워크를 설계하는 수준. 크리티컬한 과제·딜을 주도하고, 다른 구성원을 체계적으로 멘토링.",
        ),
    ]
    r = 4
    for title, desc in levels:
        ws0.cell(r, 1, title).font = Font(bold=True)
        ws0.cell(r, 2, desc).alignment = wrap
        ws0.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1

    r += 1
    ws0.cell(r, 1, "【허용 점수】0.2 단위 (종합점수 입력 시 선택)").font = hdr_font
    r += 1
    ws0.cell(r, 1, "허용점수").font = Font(bold=True)
    scores = []
    v = 1.0
    while v <= 5.001:
        scores.append(round(v, 1))
        v += 0.2
    start_score_row = r + 1
    for i, s in enumerate(scores):
        ws0.cell(start_score_row + i, 1, s)
    end_score_row = start_score_row + len(scores) - 1

    r = end_score_row + 2
    ws0.cell(r, 1, "【집계 규칙】").font = hdr_font
    rules = [
        "· 입력란 표시: 연한 노란색(#FFF2CC)은 직접 입력·목록 선택. 연한 파란색(#DDEBF7)은 「소수_누적_팀장」C~H처럼 자체평가에서 자동으로 가져오는 수식(다르게 쓰려면 수식 삭제 후 입력).",
        "· 월간 시트: 평가월은 YYYY-MM 형식으로 입력 (예: 2026-01).",
        "· 분기 열: 평가월을 기준으로 자동 산출 (예: 2026-01 → 2026-Q1).",
        "· 분기 평균: 해당 분기에 속한 월간 종합점수의 산술평균.",
        "· 분기 집계: 동일 연도 기준 Q1~Q4 평균과, 직전 분기 대비 증감·성장률%(Q2←Q1, Q3←Q2, Q4←Q3)을 본다.",
        "· 성장률(%): 직전 분기 평균이 0이 아닐 때 (당분기−직전분기)÷직전분기. 분모가 매우 작을 때는 해석에 주의.",
        "· 월간평가_RAW C열(종합점수): **같은 행의 평가월(A)·이름(B)**으로 「소수_누적_팀장」의 키(M열, 평가월|이름)를 찾아 L열(제안_종합)을 가져옵니다. 이름만 맞고 월이 다르면 연결되지 않습니다. 팀장 시트에 해당 월·이름 행이 있어야 값이 나옵니다.",
        "· 소수_누적_팀장: A~B는 팀장 입력. C~H는 기본적으로 「소수_누적_자체평가」동일 키에서 자동 참조 후, 누적·캡·MROUND·제안_종합(L)은 이 시트에서 재계산. 자체평가 시트와 구조·우선순위 동일.",
        "· 소수 누적 우선순위(두 시트 동일): ①자율성 ②범위·난이도 ③협업·영향 ④품질·완결성 ⑤전문성·표준화. 선행이 모두 충족일 때만 다음 +0.2(합산 최대 1.0).",
        "· 분기_성장률차트 시트: 분기집계의 성장률%(G,I,K)를 이름별 막대 차트로 표시(차트 데이터는 같은 파일 내 연동).",
    ]
    for line in rules:
        ws0.cell(r + 1, 1, line)
        ws0.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=6)
        ws0.cell(r + 1, 1).alignment = wrap
        r += 1

    ws0.column_dimensions["A"].width = 14
    ws0.column_dimensions["B"].width = 52

    # --- 루브릭 (행 순서 = 소수 누적 우선순위와 동일) ---
    ws_r = wb.create_sheet("루브릭")
    ws_r["A1"] = (
        "관찰 지표 초안 (차원 × 큰 레벨) — 행 순서는 「소수_누적_자체평가」「소수_누적_팀장」과 동일 "
        "(자율성 → 범위·난이도 → 협업·영향 → 품질·완결성 → 전문성·표준화)"
    )
    ws_r.merge_cells("A1:F1")
    ws_r["A1"].font = Font(bold=True, size=12)
    ws_r["A1"].alignment = wrap

    headers = ["차원", "Level 1", "Level 2", "Level 3", "Level 4", "Level 5"]
    for c, h in enumerate(headers, 1):
        cell = ws_r.cell(3, c, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border_all
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # 순서: 자율성 → 범위·난이도 → 협업·영향 → 품질·완결성 → 전문성·표준화
    rubric_rows = [
        (
            "자율성",
            "매뉴얼 없이 동일 유형도 어렵다. 일정·우선순위 대부분 배정.",
            "정형 절차 내에서는 끝까지 수행. 예외 시 질문·에스컬레이션 다수.",
            "일반 케이스는 계획~산출 스스로. 복잡 이슈는 상위와 짝.",
            "복잡 케이스도 안정적으로 끝냄. 리스크 사전 공유.",
            "영역 표준을 세우고 주도. 크리티컬 과제 리드.",
        ),
        (
            "범위·난이도",
            "단순·보조 업무 위주.",
            "단순·반복 중심. 난이도 상승 시 지원 필요.",
            "일반 난이도 담당. 전략 과제는 분리.",
            "고난이도·예외 비중 높아도 처리.",
            "가장 어려운 딜·과제 배정의 중심.",
        ),
        (
            "협업·영향",
            "협업은 지시 단위로 진행.",
            "협업 가능하나 조율은 리드가 많이 함.",
            "이해관계자와 일상 조율 가능.",
            "타 직무/고객과 선제 조율·코칭 시작.",
            "조직 단위 멘토링·정렬. 프레임으로 영향 확대.",
        ),
        (
            "품질·완결성",
            "산출물 매번 상세 리뷰·수정 필요.",
            "통과 수준. 재작업은 가끔 발생.",
            "리뷰는 방향·리스크 중심. 문장 단위 수정 감소.",
            "재작업·클레임 적음. 품질 일관.",
            "품질이 레퍼런스. 검수 기준·체크리스트 기여.",
        ),
        (
            "전문성·표준화",
            "표준 문서 숙지 단계.",
            "템플릿·가이드 준수.",
            "가이드 개선 제안 가끔.",
            "프로세스 개선·내부 가이드 초안.",
            "기준·프레임워크 설계. 레퍼런스 역할.",
        ),
    ]

    for ri, row_data in enumerate(rubric_rows, 4):
        for ci, val in enumerate(row_data, 1):
            c = ws_r.cell(ri, ci, val)
            c.alignment = wrap
            c.border = border_all
    for col in range(1, 7):
        ws_r.column_dimensions[get_column_letter(col)].width = 28 if col > 1 else 16

    # --- 소수_누적: 자체평가 / 팀장 평가 시트 분리 (월간 RAW는 팀장 시트만 참조) ---
    frac_samples = [
        ("2026-01", "김윤형", 2, 2, ("충족", "충족", "미충족", "충족", "미충족"), ("충족", "충족", "미충족", "충족", "미충족")),
        ("2026-02", "김윤형", 2, 2, ("충족", "충족", "충족", "충족", "미충족"), ("충족", "충족", "충족", "충족", "미충족")),
        ("2026-03", "김윤형", 2, 2, ("충족", "충족", "충족", "충족", "충족"), ("충족", "충족", "충족", "충족", "충족")),
        ("2026-01", "신혜윤", 3, 3, ("충족", "충족", "충족", "충족", "미충족"), ("충족", "미충족", "미충족", "미충족", "미충족")),
        ("2026-02", "신혜윤", 3, 3, ("충족", "충족", "충족", "충족", "미충족"), ("충족", "충족", "미충족", "충족", "미충족")),
        ("2026-03", "신혜윤", 3, 3, ("충족", "충족", "충족", "충족", "충족"), ("충족", "충족", "충족", "미충족", "미충족")),
        ("2026-04", "김윤형", 2, 2, ("충족", "충족", "충족", "충족", "충족"), ("충족", "충족", "충족", "충족", "미충족")),
        ("2026-04", "신혜윤", 3, 3, ("충족", "충족", "충족", "충족", "미충족"), ("충족", "충족", "충족", "충족", "미충족")),
        ("2026-07", "김윤형", 3, 2, ("충족", "충족", "충족", "충족", "충족"), ("충족", "충족", "충족", "충족", "충족")),
        ("2026-07", "신혜윤", 3, 3, ("충족", "충족", "충족", "충족", "충족"), ("충족", "충족", "충족", "충족", "미충족")),
        ("2026-07", "최우성", 2, 2, ("충족", "충족", "미충족", "충족", "미충족"), ("충족", "충족", "미충족", "충족", "미충족")),
        ("2026-10", "김윤형", 2, 2, ("충족", "충족", "충족", "충족", "미충족"), ("충족", "충족", "충족", "충족", "미충족")),
        ("2026-10", "신혜윤", 3, 3, ("충족", "충족", "충족", "충족", "충족"), ("충족", "충족", "충족", "충족", "충족")),
        ("2026-10", "최우성", 2, 2, ("충족", "충족", "충족", "충족", "미충족"), ("충족", "충족", "충족", "충족", "미충족")),
        ("2026-01", "최우성", 2, 2, ("충족", "충족", "미충족", "충족", "미충족"), ("충족", "충족", "미충족", "충족", "미충족")),
        ("2026-02", "최우성", 2, 2, ("충족", "충족", "충족", "충족", "미충족"), ("충족", "충족", "충족", "충족", "미충족")),
        ("2026-03", "최우성", 2, 2, ("충족", "충족", "충족", "충족", "충족"), ("충족", "충족", "충족", "충족", "충족")),
        ("2026-04", "최우성", 2, 2, ("충족", "충족", "충족", "충족", "충족"), ("충족", "충족", "충족", "충족", "미충족")),
    ]
    self_rows = [(ym, name, ints, *st) for ym, name, ints, intm, st, mg in frac_samples]
    ws_self = wb.create_sheet("소수_누적_자체평가")
    build_fraction_accumulation_sheet(
        ws_self,
        "소수 누적 — 자체 평가 (팀원 입력)",
        "정수레벨(C)과 5차원(D~H)은 본인이 입력. 누적·품질 캡·MROUND·제안_종합(L)은 자동. "
        "월간평가_RAW와 연결하지 않으며, 팀장 시트와 같은 평가월·이름 행을 맞추면 비교하기 쉽습니다.",
        self_rows,
    )

    ws_mgr = wb.create_sheet("소수_누적_팀장")
    build_fraction_accumulation_sheet(
        ws_mgr,
        "소수 누적 — 팀장 평가 (팀장 입력)",
        "평가월·이름(A~B)은 팀장이 입력. 정수·5차원(C~H)은 기본적으로 「소수_누적_자체평가」와 동일 키(평가월|이름) 행에서 자동으로 가져옵니다. "
        "팀장 판단이 다르면 해당 셀의 수식을 지우고 값을 직접 입력하면 됩니다. "
        "월간평가_RAW C열(종합점수)은 이 시트의 L·M열을 참조합니다.",
        self_rows,
        link_self_sheet="소수_누적_자체평가",
    )

    # --- 월간평가_RAW (C = 소수_누적_팀장!L, 키 = M) ---
    ws_m = wb.create_sheet("월간평가_RAW")
    mh = ["평가월", "이름", "종합점수", "근거요약", "분기(자동)"]
    for c, h in enumerate(mh, 1):
        cell = ws_m.cell(1, c, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border_all

    samples = [
        ("2026-01", "김윤형", "Q1 샘플"),
        ("2026-02", "김윤형", "Q1 진행"),
        ("2026-03", "김윤형", "Q1 마감"),
        ("2026-01", "신혜윤", "Q1 시작"),
        ("2026-02", "신혜윤", "Q1 중간"),
        ("2026-03", "신혜윤", "Q1 종료"),
        ("2026-04", "김윤형", "Q2"),
        ("2026-04", "신혜윤", "Q2"),
        ("2026-07", "김윤형", "Q3"),
        ("2026-07", "신혜윤", "Q3"),
        ("2026-07", "최우성", "Q3"),
        ("2026-10", "김윤형", "Q4"),
        ("2026-10", "신혜윤", "Q4"),
        ("2026-10", "최우성", "Q4"),
        ("2026-01", "최우성", "Q1"),
        ("2026-02", "최우성", "Q1"),
        ("2026-03", "최우성", "Q1"),
        ("2026-04", "최우성", "Q2"),
    ]
    for ri, (ym, name, note) in enumerate(samples, 2):
        ws_m.cell(ri, 1, ym)
        ws_m.cell(ri, 2, name)
        ws_m.cell(
            ri,
            3,
            f'=IF(OR(A{ri}="",B{ri}=""),"",IFERROR(SUMIFS(소수_누적_팀장!$L:$L,소수_누적_팀장!$M:$M,$A{ri}&"|"&$B{ri}),""))',
        )
        ws_m.cell(ri, 4, note)
        ws_m.cell(
            ri,
            5,
            f'=IF(A{ri}="","",YEAR(DATEVALUE(A{ri}&"-01"))&"-Q"&INT((MONTH(DATEVALUE(A{ri}&"-01"))-1)/3)+1)',
        )

    for col, w in zip(range(1, 6), [12, 10, 10, 36, 12]):
        ws_m.column_dimensions[get_column_letter(col)].width = w

    paint_rect_fill(ws_m, 2, 500, 1, 2, INPUT_FILL)
    paint_rect_fill(ws_m, 2, 500, 4, 4, INPUT_FILL)

    # --- 분기집계 (Q1~Q4 + 직전 분기 대비 증감·성장률) ---
    ws_q = wb.create_sheet("분기집계")
    qh = [
        "이름",
        "2026-Q1 평균",
        "2026-Q2 평균",
        "2026-Q3 평균",
        "2026-Q4 평균",
        "Q2−Q1 증감",
        "성장률%(←Q1)",
        "Q3−Q2 증감",
        "성장률%(←Q2)",
        "Q4−Q3 증감",
        "성장률%(←Q3)",
    ]
    for c, h in enumerate(qh, 1):
        cell = ws_q.cell(1, c, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border_all
        cell.alignment = Alignment(wrap_text=True)

    for ri, name in enumerate(["김윤형", "신혜윤", "최우성"], 2):
        ws_q.cell(ri, 1, name)
        ws_q.cell(
            ri,
            2,
            f'=IFERROR(AVERAGEIFS(월간평가_RAW!$C:$C,월간평가_RAW!$B:$B,$A{ri},월간평가_RAW!$E:$E,"2026-Q1"),"")',
        )
        ws_q.cell(
            ri,
            3,
            f'=IFERROR(AVERAGEIFS(월간평가_RAW!$C:$C,월간평가_RAW!$B:$B,$A{ri},월간평가_RAW!$E:$E,"2026-Q2"),"")',
        )
        ws_q.cell(
            ri,
            4,
            f'=IFERROR(AVERAGEIFS(월간평가_RAW!$C:$C,월간평가_RAW!$B:$B,$A{ri},월간평가_RAW!$E:$E,"2026-Q3"),"")',
        )
        ws_q.cell(
            ri,
            5,
            f'=IFERROR(AVERAGEIFS(월간평가_RAW!$C:$C,월간평가_RAW!$B:$B,$A{ri},월간평가_RAW!$E:$E,"2026-Q4"),"")',
        )
        ws_q.cell(ri, 6, f'=IF(AND(ISNUMBER(B{ri}),ISNUMBER(C{ri})),C{ri}-B{ri},"")')
        ws_q.cell(
            ri,
            7,
            f'=IF(AND(ISNUMBER(B{ri}),B{ri}<>0,ISNUMBER(C{ri})),(C{ri}-B{ri})/B{ri},"")',
        )
        ws_q.cell(ri, 8, f'=IF(AND(ISNUMBER(C{ri}),ISNUMBER(D{ri})),D{ri}-C{ri},"")')
        ws_q.cell(
            ri,
            9,
            f'=IF(AND(ISNUMBER(C{ri}),C{ri}<>0,ISNUMBER(D{ri})),(D{ri}-C{ri})/C{ri},"")',
        )
        ws_q.cell(ri, 10, f'=IF(AND(ISNUMBER(D{ri}),ISNUMBER(E{ri})),E{ri}-D{ri},"")')
        ws_q.cell(
            ri,
            11,
            f'=IF(AND(ISNUMBER(D{ri}),D{ri}<>0,ISNUMBER(E{ri})),(E{ri}-D{ri})/D{ri},"")',
        )
    note_row = 55
    ws_q.cell(note_row, 1, "※ 이름 행을 2행부터 아래로 복사해 팀원 추가. 연도·분기 문자열은 열 제목·수식에서 함께 변경. 안내 아래 여백은 차트·데이터 행 확보용.")
    ws_q.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=11)
    ws_q.cell(note_row, 1).alignment = wrap
    ws_q.cell(note_row, 1).font = Font(italic=True, color="666666")

    q_widths = [12, 12, 12, 12, 12, 12, 14, 12, 14, 12, 14]
    for col, w in enumerate(q_widths, 1):
        ws_q.column_dimensions[get_column_letter(col)].width = w

    paint_rect_fill(ws_q, 2, 54, 1, 1, INPUT_FILL)

    score_fmt = "0.0"
    for row in range(2, 2 + len(samples)):
        ws_m.cell(row, 3).number_format = score_fmt
    for row in range(2, 5):
        for col in range(2, 12):
            if col in (2, 3, 4, 5, 6, 8, 10):
                ws_q.cell(row, col).number_format = "0.0"
            else:
                ws_q.cell(row, col).number_format = "0.0%"

    # --- 분기_성장률차트 (분기집계 G,I,K → 막대 차트) ---
    CHART_DATA_END = 32
    ws_c = wb.create_sheet("분기_성장률차트")
    ws_c.merge_cells("A1:D1")
    ws_c["A1"] = "개인별 전분기 대비 성장률 — 차트용 데이터(분기집계와 동일 행 번호)"
    ws_c["A1"].font = Font(bold=True, size=12)
    ws_c["A1"].alignment = wrap
    ch_hdr = ["이름", "성장률(Q2←Q1)", "성장률(Q3←Q2)", "성장률(Q4←Q3)"]
    for c, h in enumerate(ch_hdr, 1):
        cell = ws_c.cell(2, c, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border_all
    for ri in range(3, CHART_DATA_END + 1):
        ws_c.cell(ri, 1, f'=IF(분기집계!A{ri}="","",분기집계!A{ri})')
        ws_c.cell(ri, 2, f'=IF(분기집계!A{ri}="","",분기집계!G{ri})')
        ws_c.cell(ri, 3, f'=IF(분기집계!A{ri}="","",분기집계!I{ri})')
        ws_c.cell(ri, 4, f'=IF(분기집계!A{ri}="","",분기집계!K{ri})')
        for col in (2, 3, 4):
            ws_c.cell(ri, col).number_format = "0.0%"
    for col, w in zip(range(1, 5), [14, 16, 16, 16]):
        ws_c.column_dimensions[get_column_letter(col)].width = w

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.style = 10
    chart.title = "개인별 전분기 대비 성장률 (%)"
    chart.y_axis.title = "성장률"
    chart.x_axis.title = "이름"
    chart.legend.position = "b"
    chart.height = 15
    chart.width = 22
    data_ref = Reference(ws_c, min_col=2, min_row=2, max_col=4, max_row=CHART_DATA_END)
    cats_ref = Reference(ws_c, min_col=1, min_row=3, max_row=CHART_DATA_END)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws_c.add_chart(chart, "F2")

    note_c = CHART_DATA_END + 2
    ws_c.merge_cells(start_row=note_c, start_column=1, end_row=note_c, end_column=8)
    ws_c.cell(
        note_c,
        1,
        "※ 차트는 위 A2:D32 구간을 사용합니다. 팀원이 30명을 넘기면 행·차트 데이터 범위를 함께 늘리고, "
        "차트를 선택한 뒤 ‘차트 디자인’에서 데이터 범위를 조정하세요. 빈 이름 행은 표시에서 제외됩니다.",
    )
    ws_c.cell(note_c, 1).alignment = wrap
    ws_c.cell(note_c, 1).font = Font(italic=True, color="666666")
    ws_c.freeze_panes = "A3"

    ws_m.freeze_panes = "A2"

    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
